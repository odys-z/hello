import os
import re
import json
import datetime
from types import SimpleNamespace

import docutils.core
from docutils.parsers.rst import Directive
from docutils.parsers.rst import directives

from sphinx.util import logging
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

logger = logging.getLogger(__name__)

def _get_resource_dir(folder):
    current_path = os.path.dirname(__file__)
    resource_path = os.path.join(current_path, folder)
    return resource_path

def az_to_dec(s):
    if not s:
        return 0

    n = 0
    i = len(s) -1
    j = 1

    while True:
        if i < 0:
            break
        c = s[i].upper()
        n += (ord(c) % 32) * j
        i -= 1
        j *= 26

    return n

# f-stringify any string
def fstr(template, **kwargs):
    return eval(f"f'{template}'", kwargs)

def processValue(row, col, cell, xforms):
    cell_value = cell.value

    if not cell_value or not xforms:
        return cell_value

    for xform in xforms:
        matcher = xform["_matcher"]
        if type(cell_value) == str and matcher.search(cell_value):
            match = matcher.search(cell_value)
            replacer = xform["replacer"]

            if xform["type"] == "decoder":
                for group_decoders in xform["group_decoders"]:
                    grp_name = group_decoders["group"]
                    decoders = group_decoders["decoders"]
                    grp_val = match.group(grp_name)
                    if decoders and grp_val and grp_val in decoders:
                        decoder = SimpleNamespace(**decoders[grp_val])
                        group = SimpleNamespace(**match.groupdict())
                        replacer = fstr(replacer, group=group, decoder=decoder)

            cell_value = matcher.sub(replacer, cell_value)

    return cell_value

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime.date):
            date_string = obj.strftime('%Y-%m-%d')
            return date_string
        if isinstance(obj, datetime.datetime):
            datetime_string = obj.strftime("%Y-%m-%d %H:%M:%S")
            return datetime_string
        return json.JSONEncoder.default(self, obj)

def _dumps(data):
    return json.dumps(data, cls=DateTimeEncoder)

class ExcelTable(Directive):

    option_spec = {
        'file': directives.path,
        'sheet': directives.unchanged,
        'rows': directives.unchanged,
        'selection': directives.unchanged,
        'overflow': directives.unchanged,
        'tablewidth': directives.unchanged,
        'colwidths': directives.unchanged,
        'row_header': directives.unchanged,
        'col_header': directives.unchanged,
        'transforms': directives.path,
    }

    def run(self, icnt=[0]):
        env = self.state.document.settings.env
        document = self.state.document

        # Track the number of times the directive is called,
        # send this to jinja template
        icnt[0] += 1

        file_path = self.options.get('file')
        selection = self.options.get('selection')
        rows = self.options.get('rows')
        sheet_name = self.options.get('sheet')
        overflow = self.options.get('overflow', 'horizontal')
        tablewidth = self.options.get('tablewidth', 'undefined')
        colwidths = self.options.get('colwidths', 'undefined')
        row_header = self.options.get('row_header', 'true')
        col_header = self.options.get('col_header', 'true')
        transforms_path = self.options.get('transforms')

        if not file_path:
            msg = "file option is missing"
            return [document.reporter.warning(msg, line=self.lineno)]

        if selection and ':' not in selection:
            msg = "selection must contain a range seperated by :"
            return [document.reporter.warning(msg, line=self.lineno)]

        if rows and ':' not in rows:
            msg = "rows must contain a range seperated by :"
            return [document.reporter.warning(msg, line=self.lineno)]

        transforms = []
        if transforms_path:
            relfn, json_mappings_file = env.relfn2path(transforms_path)
            env.note_dependency(relfn)
            f = open(json_mappings_file,)
            transformations = json.load(f)
            transforms = transformations["transforms"]
            #logger.info("transformations="+str(type(transforms[0]))+"\n")


        relfn, excel_file = env.relfn2path(file_path)
        env.note_dependency(relfn)

        if overflow != 'false':
            overflow = "\'%s\'" % overflow

        data = {
          'file_name': file_path,
          'sheet_name': sheet_name,
          'overflow': overflow,
          'tablewidth': tablewidth,
          'colwidths': colwidths,
          'row_header': row_header,
          'col_header': col_header,
          'icnt': icnt[0],
        }

        wb = load_workbook(filename=excel_file)

        if sheet_name and sheet_name not in wb.sheetnames:
            msg = "sheet %s does not exist" % sheet_name
            return [document.reporter.warning(msg, line=self.lineno)]

        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.worksheets[0]

        if selection:
            sheet_data = sheet[selection]
        elif rows:
            sheet_data = sheet[rows]
        else:
            sheet_data = sheet

        # setup transform cell array based on rows and colums in sheet_data.
        xforms = [[[None] * len(transforms)] * len(sheet_data[0])] * len(sheet_data)
        if transforms:
            # pre-compile all matcher and replace regexs
            for t, transform in enumerate(transforms):
                # logger.info( transform["matcher"])
                if "matcher" not in transform:
                    msg = "configured transformations must have a 'matcher' element."
                    return [document.reporter.warning(msg, line=self.lineno)]
                transform["_matcher"] = re.compile( transform["matcher"])
                transform["_rows"] = []
                if "rows" in transform:
                    transform["_rows"] = [int(x) - 1 for x in transform["rows"].split(',') if x.strip().isdigit()]

                transform["_cols"] = []
                if "cols" in transform:
                    transform["_cols"] = [int(x) - 1 for x in transform["cols"].split(',') if x.strip().isdigit()]

                rows = range(len(sheet_data))
                cols = range(len(sheet_data[0]))
                if transform["_rows"]:
                    rows = transform["_rows"]
                if transform["_cols"]:
                    cols = transform["_cols"]
                for r in rows:
                    for c in cols:
                        xforms[r][c][t] = transform

        content = []
        for rownum, row in enumerate(sheet_data):
            _row = {}
            for colnum, cell in enumerate(row):
                _row[colnum] = processValue(rownum, colnum, cell, xforms[rownum][colnum])
            content.append(_row)

        data['content'] = _dumps(content)

        columns = []
        for i, cell in enumerate(sheet_data[0]):
            _col = {}
            _col["data"] = i
            _col["renderer"] = "html"
            columns.append(_col)

        data['columns'] = _dumps(columns)

        cell_regex = re.compile(r'([A-Z]+){1}([0-9]+){1}')
        merged_cells = []

        for m in sheet.merged_cells:
            start, end = str(m).split(":")

            start_match = cell_regex.search(start)
            start_col, start_row = start_match.groups()

            end_match = cell_regex.search(end)
            end_col, end_row = end_match.groups()

            m_row = int(start_row) - 1
            m_col = az_to_dec(start_col) - 1
            m_rowspan = int(end_row) - int(start_row) + 1
            m_colspan = az_to_dec(end_col) - az_to_dec(start_col) + 1

            merged_cells.append({
                "row": m_row,
                "col": m_col,
                "rowspan": m_rowspan,
                "colspan": m_colspan,
            })

        data['merged_cells'] = _dumps(merged_cells)

        loader = FileSystemLoader(_get_resource_dir('templates'))
        _env = Environment(loader=loader,
                           keep_trailing_newline=True,
                           trim_blocks=True,
                           lstrip_blocks=True)

        template = _env.get_template('table.html')
        html = template.render(**data)
        return [docutils.nodes.raw('', html, format='html')]
